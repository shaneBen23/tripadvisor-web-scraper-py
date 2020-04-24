from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

def scrapeAddress(url):
  chrome_options = webdriver.ChromeOptions()
  chrome_options.add_argument("--headless")

  browser = webdriver.Chrome(options=chrome_options)
  browser.get(url)
  
  address = "N/A"

  try:
    elem = browser.find_element(By.XPATH, "//*[@id='taplc_resp_rr_top_info_rr_resp_0']/div/div[4]/div[1]/div/div/div[1]/span[2]")
    address = elem.text
  except:
    print("Page not available")

  browser.quit()

  return address

def updatedLondonListingsAddress():
  filename = "london_listings_updated.xlsx"
  baseURL = "https://www.tripadvisor.co.uk"
  DONE = "done"
  workbook = load_workbook(filename=filename)
  sheet = workbook.active

  rows = sheet.max_row + 1

  for i in range(2, rows):
    name = sheet[f"B{i}"]
    path = sheet[f"J{i}"]
    cell = sheet[f"L{i}"]
    done = sheet[f"M{i}"]

    print("Row number:", i)
    print("Restaurant name:", name.value)

    if cell.value:
      print("Restaurant address:", cell.value)
    
    if done.value != DONE:
      address = scrapeAddress(baseURL + path.value)
      sheet[f"L{i}"] = address
      sheet[f"M{i}"] = DONE

      print("Restaurant address:", cell.value)
      if i % 10 == 0:
        print("****** Writing to file ******")
        workbook.save(filename=filename)
        print("****** Writing complete ******")
      
      if i % 100 == 0:
        print("****** Writing backup ******")
        workbook.save(filename="london_listings_backup.xlsx")
        print("****** Backup complete ******")
      
      print("===")
    else:
      print("****** Already up to date ******")
      print("===")

  print('********************************')
  print('Listing address update completed')
  print('********************************')

def main():
  updatedLondonListingsAddress()

if __name__ == "__main__":
  main()
